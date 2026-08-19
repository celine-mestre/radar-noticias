#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Radar de Notícias por Área Governativa — extração para Excel
Secretaria-Geral do Governo · Suporte à Decisão · Unidade de Pesquisa e Estatísticas

Lê os feeds das 16 áreas governativas e grava um ficheiro Excel com uma linha por
notícia: área, agrupamento, data de publicação, fonte, título, resumo e ligação.

Utilização:
    python extrair_noticias.py                      janela predefinida (7 dias)
    python extrair_noticias.py --periodo 24h        últimas 24 horas
    python extrair_noticias.py --periodo 30d --saida noticias_julho.xlsx
    python extrair_noticias.py --area saude         apenas uma área

Requisitos: Python 3.9 ou superior e a biblioteca openpyxl (pip install openpyxl).
"""

import argparse
import html
import json
import unicodedata
import os
import re
import sys
import time
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

# ── HORAS ────────────────────────────────────────────────────────────────────
# Tudo o que este programa escreve fica na hora de Lisboa. São dois problemas
# distintos e ambos davam incoerências visíveis no painel:
#
#  1. O servidor do GitHub corre em UTC. Escrevendo agora_lisboa(), a hora da
#     recolha saía uma hora atrás da real no horário de verão.
#  2. Os feeds datam os artigos no seu próprio fuso. Retirar o fuso sem
#     converter — que era o que se fazia — deixava a hora de Berlim ou de
#     Bruxelas como se fosse a nossa, e apareciam notícias "do futuro".
#
# Converte-se tudo para Europa/Lisboa e só depois se retira o fuso, para que as
# comparações e o que se apresenta digam respeito ao mesmo relógio.
LISBOA = ZoneInfo("Europe/Lisbon")


def agora_lisboa():
    """Hora de Lisboa, sem fuso, para gravar e comparar."""
    return datetime.now(LISBOA).replace(tzinfo=None)


def para_lisboa(momento):
    """Converte uma data com fuso para hora de Lisboa, sem fuso."""
    if momento is None:
        return None
    if momento.tzinfo is None:
        return momento
    return momento.astimezone(LISBOA).replace(tzinfo=None)


from email.utils import parsedate_to_datetime
from html.entities import name2codepoint
from base64 import urlsafe_b64decode
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import quote, urlparse
from urllib.request import Request, urlopen
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CONFIGURAÇÃO — manter alinhada com o painel HTML
# ---------------------------------------------------------------------------
EDICAO = {"hl": "pt-PT", "gl": "PT", "ceid": "PT:pt-150"}
# Domínios excluídos quando se restringe a fontes nacionais: domínios de topo dos
# restantes países de língua portuguesa e grupos de comunicação social que usam
# domínios genéricos.
EXCLUSOES_DOMINIO = [
    "site:.br", "site:.ao", "site:.mz", "site:.cv", "site:.st", "site:.gw",
    "site:.tl", "site:.mo",
    "site:globo.com", "site:r7.com", "site:metropoles.com", "site:abril.com",
    "site:uol.com.br", "site:novojornal.co.ao", "site:jornaldeangola.ao",
    "site:verangola.net", "site:opais.co.mz", "site:verdade.co.mz",
    "site:expressodasilhas.cv", "site:asemana.publ.cv",
]

GRUPOS = {
    "soberania": "Estado e soberania",
    "economia": "Economia, finanças e território",
    "social": "Sociedade e bem-estar",
    "ambiente": "Ambiente, energia e recursos naturais",
}

AREAS = [
    # O Primeiro-Ministro encabeça a lista: precede protocolarmente todas as
    # áreas ministeriais. O cargo em absoluto ("primeiro-ministro") não entra
    # nas palavras — apanharia os chefes de governo estrangeiros, e enumerar
    # qualificativos nunca chega (testado: Bulgária, Israel e Hungria escapavam).
    # Fica o apelido, que a imprensa portuguesa usa quase sempre, com exclusões
    # locativas para o país homónimo e para a cidade gaúcha de Montenegro (RS).
    # A sigla "PM" também fica de fora (Polícia Marítima, "3 pm").
    ("primeiro_ministro", "Primeiro-Ministro", "soberania",
     [
      "Luís Montenegro",
      "Montenegro",
      "primeiro-ministro português",
      "chefe do Governo português",
      "XXV Governo"],
     [
      "no Montenegro",
      "ao Montenegro",
      "pelo Montenegro",
      "em Montenegro",
      "cidade de Montenegro",
      "município de Montenegro",
      "Montenegro e Sérvia",
      "Sérvia e Montenegro",
      "Montenegro, Sérvia",
      "capital do Montenegro",
      "Podgorica",
      "Bósnia",
      "ex-primeiro-ministro",
      "antigo primeiro-ministro"]),
    ("estrangeiros", "Negócios Estrangeiros", "soberania",
     [
      "Ministério dos Negócios Estrangeiros",
      "política externa",
      "diplomacia",
      "assuntos europeus",
      "Conselho Europeu",
      "União Europeia",
      "CPLP",
      "Nações Unidas",
      "comunidades portuguesas",
      "emigrantes",
      "rede consular",
      "cooperação para o desenvolvimento",
      "ministro dos Negócios Estrangeiros",
      "secretário de Estado dos Assuntos Europeus",
      "secretário de Estado dos Negócios Estrangeiros",
      "secretário de Estado das Comunidades Portuguesas"],
     []),
    ("financas", "Finanças", "economia",
     [
      "Ministério das Finanças",
      "Orçamento do Estado",
      "défice",
      "dívida pública",
      "impostos",
      "IRS",
      "IRC",
      "IVA",
      "carga fiscal",
      "Autoridade Tributária",
      "Administração Pública",
      "funcionários públicos",
      "execução orçamental",
      "ministro das Finanças",
      "secretário de Estado Adjunto e do Orçamento",
      "secretário de Estado dos Assuntos Fiscais",
      "secretário de Estado do Tesouro",
      "secretário de Estado da Administração Pública"],
     []),
    ("presidencia", "Presidência", "soberania",
     [
      "Ministério da Presidência",
      "Conselho de Ministros",
      "Presidência do Conselho de Ministros",
      "Programa do Governo",
      "comunicação do Governo",
      "imigração",
      "imigrantes",
      "migrantes",
      "migratória",
      "AIMA",
      "autorização de residência",
      "vistos de residência",
      "nacionalidade portuguesa",
      "regularização de estrangeiros",
      "acolhimento e integração",
      "ministro da Presidência",
      "secretário de Estado Adjunto da Presidência",
      "secretário de Estado da Presidência"],
     []),
    ("reforma", "Reforma do Estado", "soberania",
     [
      "reforma do Estado",
      "simplificação administrativa",
      "desburocratização",
      "transformação digital",
      "serviços públicos digitais",
      "automatização de processos",
      "interoperabilidade",
      "identificação digital",
      "dados abertos",
      "inteligência artificial",
      "modernização administrativa",
      "SIADAP",
      "avaliação de desempenho",
      "ministro Adjunto e da Reforma do Estado",
      "ministro da Reforma do Estado",
      "secretário de Estado para a Digitalização",
      "secretário de Estado para a Simplificação"],
     ["inteligência artificial generativa em contexto militar"]),
    ("parlamentares", "Assuntos Parlamentares", "soberania",
     [
      "Assembleia da República",
      "debate parlamentar",
      "interpelação ao Governo",
      "audição parlamentar",
      "comissão de inquérito",
      "iniciativa legislativa",
      "moção de censura",
      "regulação da comunicação social",
      "ministro dos Assuntos Parlamentares"],
     []),
    ("defesa", "Defesa Nacional", "soberania",
     [
      "Ministério da Defesa Nacional",
      "Forças Armadas",
      "militares",
      "NATO",
      "política de defesa",
      "investimento em defesa",
      "indústria de defesa",
      "missões militares",
      "serviço militar",
      "Ucrânia",
      "ministro da Defesa",
      "secretário de Estado Adjunto e da Defesa Nacional",
      "secretário de Estado da Defesa"],
     ["defesa do consumidor", "defesa pessoal", "linha de defesa", "defesa do título"]),
    ("interna", "Administração Interna", "soberania",
     [
      "Ministério da Administração Interna",
      "segurança interna",
      "forças de segurança",
      "PSP",
      "GNR",
      "fronteiras",
      "asilo",
      "refugiados",
      "deportação",
      "expulsão do território",
      "proteção civil",
      "incêndios florestais",
      "sinistralidade rodoviária",
      "ministro da Administração Interna",
      "secretário de Estado da Administração Interna",
      "secretário de Estado Adjunto da Administração Interna",
      "secretário de Estado da Proteção Civil"],
     ["fronteira entre", "fronteiras do conhecimento", "últimas fronteiras"]),
    ("justica", "Justiça", "soberania",
     [
      "Ministério da Justiça",
      "tribunais",
      "Ministério Público",
      "magistrados",
      "processos pendentes",
      "sistema prisional",
      "prisões",
      "registos e notariado",
      "reforma da justiça",
      "apoio judiciário",
      "corrupção",
      "ministro da Justiça",
      "secretário de Estado da Justiça",
      "secretário de Estado Adjunto e da Justiça"],
     ["justiça poética", "fazer justiça a"]),
    ("economia", "Economia e Coesão Territorial", "economia",
     [
      "Ministério da Economia",
      "crescimento económico",
      "exportações",
      "investimento",
      "tecido empresarial",
      "PME",
      "apoios às empresas",
      "empresas portuguesas",
      "PRR",
      "Plano de Recuperação e Resiliência",
      "turismo",
      "comércio",
      "fundos europeus",
      "desenvolvimento regional",
      "coesão territorial",
      "administração local",
      "autarquias",
      "ordenamento do território",
      "interioridade",
      "ministro da Economia",
      "secretário de Estado da Economia",
      "secretário de Estado do Turismo",
      "secretário de Estado da Administração Local",
      "secretário de Estado do Planeamento"],
     ["comércio eletrónico de dados", "ambiente de trabalho", "empresa familiar"]),
    ("infraestruturas", "Infraestruturas e Habitação", "economia",
     [
      "Ministério das Infraestruturas e Habitação",
      "habitação",
      "arrendamento",
      "preço das casas",
      "compra de casa",
      "obras públicas",
      "ferrovia",
      "comboios",
      "novo aeroporto",
      "transportes públicos",
      "mobilidade",
      "licenciamento urbanístico",
      "ministro das Infraestruturas",
      "secretário de Estado das Infraestruturas",
      "secretário de Estado da Mobilidade",
      "secretário de Estado da Habitação"],
     []),
    ("educacao", "Educação, Ciência e Inovação", "social",
     [
      "Ministério da Educação",
      "educação pré-escolar",
      "creches",
      "ensino básico",
      "ensino secundário",
      "ensino profissional",
      "escolas",
      "alunos",
      "professores",
      "exames nacionais",
      "manuais escolares",
      "abandono escolar",
      "ensino superior",
      "universidades",
      "ciência e inovação",
      "investigação científica",
      "investigação e desenvolvimento",
      "bolsas de estudo",
      "bolsas de investigação",
      "ministro da Educação",
      "secretário de Estado Adjunto e da Educação",
      "secretário de Estado da Administração Escolar",
      "secretário de Estado da Ciência e Inovação",
      "secretário de Estado do Ensino Superior"],
     ["bolsa de valores", "bolsa espanhola", "bolsa de Lisboa"]),
    ("saude", "Saúde", "social",
     [
      "Ministério da Saúde",
      "Serviço Nacional de Saúde",
      "hospitais",
      "urgências",
      "médicos de família",
      "listas de espera",
      "cuidados continuados",
      "medicamentos",
      "saúde mental",
      "enfermeiros",
      "ministro da Saúde",
      "secretário de Estado da Saúde",
      "secretário de Estado da Gestão da Saúde"],
     []),
    ("trabalho", "Trabalho, Solidariedade e Segurança Social", "social",
     [
      "Ministério do Trabalho",
      "emprego",
      "desemprego",
      "salário mínimo",
      "salários",
      "pensões",
      "reformados",
      "segurança social",
      "ação social",
      "greve",
      "negociação coletiva",
      "apoios sociais",
      "ministro do Trabalho",
      "secretário de Estado do Trabalho",
      "secretário de Estado da Segurança Social",
      "secretário de Estado da Ação Social"],
     []),
    ("cultura", "Cultura, Juventude e Desporto", "social",
     [
      "Ministério da Cultura",
      "património cultural",
      "museus",
      "bibliotecas",
      "artes performativas",
      "espetáculos",
      "cinema",
      "programação cultural",
      "artistas",
      "políticas de juventude",
      "associativismo juvenil",
      "jovens",
      "igualdade de género",
      "desporto",
      "prática desportiva",
      "federações desportivas",
      "alta competição",
      "atletas",
      "língua portuguesa",
      "ministro da Cultura",
      "secretário de Estado da Cultura",
      "secretário de Estado do Desporto",
      "secretário de Estado adjunto da Juventude",
      "secretário de Estado da Juventude"],
     ["cultura organizacional", "cultura de empresa", "cultura do cancelamento", "desporto-rei", "património genético", "património imobiliário"]),
    ("ambiente", "Ambiente e Energia", "ambiente",
     [
      "Ministério do Ambiente e Energia",
      "energia",
      "eletricidade",
      "renováveis",
      "alterações climáticas",
      "descarbonização",
      "resíduos",
      "seca",
      "abastecimento de água",
      "qualidade da água",
      "escassez de água",
      "poluição",
      "ministro do Ambiente",
      "secretário de Estado do Ambiente",
      "secretário de Estado da Energia"],
     ["ambiente de trabalho", "ambiente empresarial", "ambiente de negócios"]),
    ("agricultura", "Agricultura e Mar", "ambiente",
     [
      "Ministério da Agricultura e Mar",
      "política agrícola comum",
      "agricultura",
      "agricultores",
      "desenvolvimento rural",
      "florestas",
      "gestão florestal",
      "pescas",
      "pescadores",
      "aquicultura",
      "economia do mar",
      "segurança alimentar",
      "regadio",
      "ministro da Agricultura",
      "secretário de Estado da Agricultura",
      "secretário de Estado das Pescas",
      "secretário de Estado das Florestas"],
     [])

]

AZUL, CINZA = "2B5683", "F2F5F8"

# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Publicações recolhidas VIA GOOGLE NOTÍCIAS (pesquisa site:domínio)
#
# A verificação de agosto de 2026 mostrou dois becos sem saída: órgãos que
# bloqueiam pedidos vindos de IPs de centros de dados como os do GitHub
# (Expresso, SIC, JN, TSF respondem 403 faça-se o que se fizer ao endereço) e
# órgãos sem feed encontrável — nem nos caminhos conhecidos nem no que a
# própria página anuncia. Para estes, o feed lê-se ao Google Notícias restrito
# ao domínio da publicação, reaproveitando o circuito da era Google que o
# projeto guarda: o <source> traz o nome e o domínio do órgão, a cauda
# « - Fonte» do título é retirada, e as ligações de reencaminhamento
# resolvem-se para o endereço do jornal no fim da recolha.
#
# A janela é de um dia (when:1d) para ficar aquém do teto de 100 resultados
# por consulta — acima disso o Google escolhe por relevância e a cronologia
# deixa de ser garantida; com oito recolhas diárias, nada se perde.
# ---------------------------------------------------------------------------
VIA_GOOGLE = {
    # Bloqueiam IPs de centros de dados (403 comprovado em duas verificações)
    "Expresso", "SIC Notícias", "Jornal de Notícias", "TSF",
    # Sem feed encontrável (verificação com autodescoberta, agosto de 2026)
    "Diário de Notícias", "Renascença", "Diário de Notícias da Madeira",
    "Jornal i", "JM Madeira", "Vida Económica", "Construir",
    "Executive Digest", "Jornal de Angola", "Novo Jornal (Angola)",
    "Angop", "Inforpress (Cabo Verde)", "Deutsche Welle (português)",
    # Sem feed público (agência por assinatura); via Google, o sítio aberto
    # da Lusa volta a contar
    "Lusa",
}


def url_via_google(dominio, periodo="1d"):
    """Endereço do feed do Google Notícias restrito a uma publicação."""
    return url_feed(f"site:{dominio} when:{periodo}")


# FONTES — os feeds das próprias publicações
#
# Em vez de interrogar um motor de pesquisa, subscreve-se o feed de cada publicação
# e faz-se a marcação por palavras-chave do nosso lado.
# Não há teto de resultados, a ordenação é cronológica, as datas são as de
# publicação, as ligações são diretas e as descrições trazem o lead da redação.
# ---------------------------------------------------------------------------
FONTES = [
    ("Público", "publico.pt", "https://feeds.feedburner.com/PublicoRSS"),
    # Os feeds temáticos do FeedBurner foram descontinuados; o Público serve-os
    # hoje no próprio sítio, em publico.pt/rss/<secção>.
    # A Impresa (Expresso, SIC) publica no padrão da plataforma Arc:
    # /arc/outboundfeeds/rss/ — os caminhos /rss antigos deixaram de responder.
    ("Expresso", "expresso.pt", "https://expresso.pt/arc/outboundfeeds/rss/?outputType=xml"),
    ("Observador", "observador.pt", "https://observador.pt/feed/"),
    ("Jornal de Notícias", "jn.pt", "https://www.jn.pt/rss/"),
    ("Diário de Notícias", "dn.pt", "https://www.dn.pt/rss/"),
    ("Correio da Manhã", "cmjornal.pt", "https://www.cmjornal.pt/rss"),
    ("Jornal de Negócios", "jornaldenegocios.pt", "https://www.jornaldenegocios.pt/rss"),
    ("Jornal Económico", "jornaleconomico.pt", "https://jornaleconomico.sapo.pt/feed"),
    ("ECO", "eco.sapo.pt", "https://eco.sapo.pt/feed/"),
    ("RTP Notícias", "rtp.pt", "https://www.rtp.pt/noticias/rss"),
    ("SIC Notícias", "sicnoticias.pt", "https://sicnoticias.pt/arc/outboundfeeds/rss/?outputType=xml"),
    ("CNN Portugal", "cnnportugal.iol.pt", "https://cnnportugal.iol.pt/rss"),
    ("TSF", "tsf.pt", "https://www.tsf.pt/rss/"),
    ("Renascença", "rr.pt", "https://rr.pt/rss"),
    ("Notícias ao Minuto", "noticiasaominuto.com", "https://www.noticiasaominuto.com/rss/ultima-hora"),
    ("Diário de Notícias da Madeira", "dnoticias.pt", "https://www.dnoticias.pt/rss"),
    ("Sábado", "sabado.pt", "https://www.sabado.pt/rss"),
    ("Visão", "visao.pt", "https://visao.pt/feed/"),
    ("Dinheiro Vivo", "dinheirovivo.pt", "https://www.dinheirovivo.pt/feed"),
    ("Executive Digest", "executivedigest.sapo.pt", "https://executivedigest.sapo.pt/feed/"),
    ("Ambiente Magazine", "ambientemagazine.com", "https://www.ambientemagazine.com/feed/"),
    ("Agroportal", "agroportal.pt", "https://www.agroportal.pt/feed/"),

    # Agência nacional
    ("Lusa", "lusa.pt", "https://www.lusa.pt/rss"),

    # Generalistas e semanários
    ("Nascer do SOL", "sol.sapo.pt", "https://sol.sapo.pt/rss"),
    ("Jornal i", "ionline.pt", "https://ionline.pt/feed/"),

    # Regionais das regiões autónomas
    ("Açoriano Oriental", "acorianooriental.pt", "https://www.acorianooriental.pt/feed/rss.xml"),
    ("JM Madeira", "jm-madeira.pt", "https://www.jm-madeira.pt/rss"),

    # Especializadas, por matéria de área governativa
    ("Vida Económica", "vidaeconomica.pt", "https://www.vidaeconomica.pt/feed"),
    ("SAPO Tek", "tek.sapo.pt", "https://tek.sapo.pt/rss"),
    ("Healthnews", "healthnews.pt", "https://healthnews.pt/feed/"),
    ("Construir", "construir.pt", "https://www.construir.pt/feed/"),
]

# Imprensa dos restantes países de língua portuguesa. Matéria da CPLP, cooperação,
# diáspora e política externa é frequentemente tratada primeiro nestes títulos.
FONTES_LUSOFONAS = [
    ("Jornal de Angola", "jornaldeangola.ao", "https://www.jornaldeangola.ao/ao/rss/"),
    ("Novo Jornal (Angola)", "novojornal.co.ao", "https://www.novojornal.co.ao/rss"),
    ("Angop", "angop.ao", "https://www.angop.ao/rss/ultimas.xml"),
    ("O País (Moçambique)", "opais.co.mz", "https://opais.co.mz/feed/"),
    ("Carta de Moçambique", "cartamz.com", "https://cartamz.com/feed/"),
    ("Expresso das Ilhas (Cabo Verde)", "expressodasilhas.cv", "https://expressodasilhas.cv/rss"),
    ("Inforpress (Cabo Verde)", "inforpress.cv", "https://inforpress.cv/feed/"),
    ("Tatoli (Timor-Leste)", "tatoli.tl", "https://tatoli.tl/feed/"),
    ("STP-Press (São Tomé)", "stp-press.st", "https://www.stp-press.st/feed/"),
    ("Agência Brasil", "agenciabrasil.ebc.com.br", "https://agenciabrasil.ebc.com.br/rss/ultimasnoticias/feed.xml"),
    ("Folha de S.Paulo · Mundo", "folha.uol.com.br", "https://feeds.folha.uol.com.br/mundo/rss091.xml"),
]

# Imprensa internacional não lusófona, para ver como uma matéria portuguesa ou
# europeia é tratada fora.
FONTES_INTERNACIONAIS = [
    ("Euronews (português)", "pt.euronews.com", "https://pt.euronews.com/rss"),
    ("Politico Europe", "politico.eu", "https://www.politico.eu/feed/"),
    ("El País", "elpais.com", "https://feeds.elpais.com/mrss-s/pages/ep/site/elpais.com/portada"),
    ("El País · Internacional", "elpais.com", "https://feeds.elpais.com/mrss-s/pages/ep/site/elpais.com/section/internacional/portada"),
    ("Le Monde", "lemonde.fr", "https://www.lemonde.fr/rss/une.xml"),
    ("BBC Mundo", "bbc.com", "https://feeds.bbci.co.uk/mundo/rss.xml"),
    # O endpoint /rdf/ da DW foi descontinuado; o formato atual é /xml/.
    ("Deutsche Welle (português)", "dw.com", "https://rss.dw.com/xml/rss-br-all"),
    # A France 24 não tem edição em português (é a RFI que a tem, e já cá está);
    # fica a edição inglesa da Europa, que serve o propósito do clipping.
    ("France 24 (inglês)", "france24.com", "https://www.france24.com/en/europe/rss"),
    ("RFI (português)", "rfi.fr", "https://www.rfi.fr/pt/rss"),
    ("The Guardian · Europe", "theguardian.com", "https://www.theguardian.com/world/europe-news/rss"),

    # Reino Unido
    ("BBC News · Mundo", "bbc.co.uk", "https://feeds.bbci.co.uk/news/world/rss.xml"),
    ("The Guardian · Mundo", "theguardian.com", "https://www.theguardian.com/world/rss"),

    # França
    ("Le Figaro", "lefigaro.fr", "https://www.lefigaro.fr/rss/figaro_actualites.xml"),
    ("France Info", "francetvinfo.fr", "https://www.francetvinfo.fr/titres.rss"),

    # Espanha
    ("El Mundo", "elmundo.es", "https://e00-elmundo.uecdn.es/elmundo/rss/portada.xml"),
    ("La Vanguardia", "lavanguardia.com", "https://www.lavanguardia.com/rss/home.xml"),
    ("ABC", "abc.es", "https://www.abc.es/rss/2.0/ultima-hora/"),

    # Itália
    ("ANSA", "ansa.it", "https://www.ansa.it/sito/ansait_rss.xml"),
    ("Corriere della Sera", "corriere.it", "https://www.corriere.it/rss/homepage.xml"),
    ("La Repubblica", "repubblica.it", "https://www.repubblica.it/rss/homepage/rss2.0.xml"),

    # Estados Unidos da América
    ("The New York Times · Mundo", "nytimes.com", "https://rss.nytimes.com/services/xml/rss/nyt/World.xml"),
    ("The Washington Post · Mundo", "washingtonpost.com", "https://feeds.washingtonpost.com/rss/world"),
    ("Politico", "politico.com", "https://rss.politico.com/politics-news.xml"),

    # Alemanha
    ("Der Spiegel · Internacional", "spiegel.de", "https://www.spiegel.de/international/index.rss")
]


# ---------------------------------------------------------------------------
# Construção das consultas
# ---------------------------------------------------------------------------
def consulta(palavras, excluir, periodo, so_nacionais=True):
    termos = "(" + " OR ".join(f'"{p}"' for p in palavras) + ")"
    termos += "".join(f' -"{t}"' for t in excluir)
    if so_nacionais:
        termos += " " + " ".join(f"-{d}" for d in EXCLUSOES_DOMINIO)
    if periodo:
        termos += f" when:{periodo}"
    return termos


def url_feed(q):
    p = f"hl={EDICAO['hl']}&gl={EDICAO['gl']}&ceid={EDICAO['ceid']}"
    return f"https://news.google.com/rss/search?q={quote(q, safe='')}&{p}"


# ---------------------------------------------------------------------------
# Leitura e tratamento
# ---------------------------------------------------------------------------
def limpar(texto):
    """Remove marcação HTML e normaliza espaços."""
    if not texto:
        return ""
    texto = re.sub(r"<[^>]+>", " ", texto)
    texto = html.unescape(texto)
    return re.sub(r"\s+", " ", texto).strip()


DOMINIOS_ALHEIOS = ("google.com", "google.pt", "gstatic.com", "googleusercontent.com",
                    "googleapis.com", "policies.google", "accounts.google", "consent.google")

# Publicações dos restantes países de língua portuguesa, para referência de quem
# analisar os ficheiros. A seleção por origem é feita no painel, sobre o domínio.
PUBLICACOES_LUSOFONAS = (
    "globo.com", "uol.com.br", "r7.com", "metropoles.com", "abril.com", "terra.com.br",
    "conjur.com.br", "migalhas.com.br", "jota.info", "exame.com",
    "jornaldeangola.ao", "novojornal.co.ao", "verangola.net", "angop.ao",
    "opais.co.mz", "idolo.co.mz", "verdade.co.mz", "carta.co.mz",
    "expressodasilhas.cv", "asemana.publ.cv", "inforpress.cv", "vozdoarquipelago.com",
    "tatoli.tl", "timorpost.com", "stp-press.st", "odemocratagb.com",
)

FICHEIROS_DE_IMAGEM = re.compile(r"\.(png|jpe?g|gif|webp|svg|ico|bmp)(\?|#|$)", re.IGNORECASE)


def endereco_plausivel(url):
    """Verifica se o endereço encontrado é mesmo o artigo de um jornal.

    Sem esta verificação, a leitura da página de reencaminhamento devolve o
    primeiro endereço que lá encontra — que costuma ser uma imagem alojada nos
    servidores do próprio Google, igual para todos os artigos.
    """
    if not url or not url.lower().startswith(("http://", "https://")) or len(url) < 24:
        return False
    if any(d in url for d in DOMINIOS_ALHEIOS):
        return False
    if FICHEIROS_DE_IMAGEM.search(url):
        return False
    return True


def endereco_do_artigo(ligacao, tempo_limite=12):
    """Converte o reencaminhamento do Google no endereço do jornal.

    Primeiro tenta descodificar o identificador, onde as versões mais antigas
    trazem o endereço em claro. Não resultando, segue o reencaminhamento na rede
    e devolve o endereço final. Falhando ambos, devolve a ligação original, que
    abre o artigo à mesma — é preferível a devolver um endereço errado.
    """
    if not ligacao or "news.google.com" not in ligacao:
        return ligacao

    identificador = ligacao.split("/articles/")[-1].split("?")[0]
    try:
        bruto = urlsafe_b64decode(identificador + "=" * (-len(identificador) % 4))
        for achado in re.findall(rb"https?://[A-Za-z0-9\-._~:/?#\[\]@!$&'()*+,;=%]{20,}", bruto):
            candidato = achado.decode("utf-8", "replace")
            if endereco_plausivel(candidato):
                return candidato
    except Exception:                                          # noqa: BLE001
        pass

    try:
        pedido = Request(ligacao, headers={"User-Agent": "Mozilla/5.0 SGGov-UPE-Radar/1.0"})
        with urlopen(pedido, timeout=tempo_limite) as resposta:
            final = resposta.geturl()
            if endereco_plausivel(final):
                return final
            corpo = resposta.read(60000).decode("utf-8", "replace")
        # O atributo data-n-au traz o endereço do artigo na página de reencaminhamento
        for padrao in (r'data-n-au="([^"]+)"',
                       r'<meta[^>]+http-equiv="refresh"[^>]+url=([^">]+)',
                       r'<link[^>]+rel="canonical"[^>]+href="([^"]+)"'):
            for achado in re.findall(padrao, corpo, re.IGNORECASE):
                candidato = html.unescape(achado).strip()
                if endereco_plausivel(candidato):
                    return candidato
    except Exception:                                          # noqa: BLE001
        pass

    return ligacao


def resolver_ligacoes(linhas, trabalhadores=8):
    """Resolve em paralelo os reencaminhamentos de todas as notícias."""
    enderecos = [l[7] for l in linhas]
    unicos = list(dict.fromkeys(enderecos))
    print(f"A resolver {len(unicos)} ligações…", end=" ", flush=True)
    with ThreadPoolExecutor(max_workers=trabalhadores) as piscina:
        resolvidos = dict(zip(unicos, piscina.map(endereco_do_artigo, unicos)))
    convertidas = 0
    for l in linhas:
        novo = resolvidos.get(l[7], l[7])
        if novo != l[7]:
            convertidas += 1
            l[7] = novo
    print(f"{convertidas} convertidas em endereço do jornal")
    return linhas


def ler_feed(url, tempo_limite=30):
    # A identificação «SGGov-UPE-Radar/1.0» era recusada pelas proteções
    # anti-robô de vários órgãos (Expresso, JN, CM, TSF, entre outros): o
    # servidor respondia 403 e o feed parecia morto sem o estar. Os cabeçalhos
    # passam a ser os de um navegador corrente, com o Accept dos formatos de
    # feed — que é o que qualquer leitor de RSS envia.
    pedido = Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/126.0.0.0 Safari/537.36",
        "Accept": "application/rss+xml, application/atom+xml, "
                  "application/xml;q=0.9, text/xml;q=0.8, */*;q=0.7",
        "Accept-Language": "pt-PT,pt;q=0.9,en;q=0.6",
    })
    with urlopen(pedido, timeout=tempo_limite) as resposta:
        return resposta.read()


def descodificar(bruto):
    """Converte os bytes do feed em texto, respeitando a codificação declarada.

    Nem todas as publicações publicam em UTF-8: algumas usam ISO-8859-1 ou
    Windows-1252. Descodificar tudo como UTF-8 fazia aparecer losangos com ponto
    de interrogação no lugar dos acentos — "incêndios" saía "inc?ndios".

    A codificação vem declarada na primeira linha do XML. Não vindo, tenta-se
    UTF-8 e recorre-se ao Windows-1252, que é o mais frequente na imprensa e
    aceita qualquer byte sem falhar.
    """
    if not isinstance(bruto, bytes):
        return bruto

    declarada = None
    achado = re.search(rb'encoding=["\']([\w-]+)["\']', bruto[:200], re.I)
    if achado:
        declarada = achado.group(1).decode("ascii", "ignore").lower()

    tentativas = []
    if declarada:
        tentativas.append(declarada)
    tentativas += ["utf-8", "cp1252", "iso-8859-1"]

    for codificacao in tentativas:
        try:
            return bruto.decode(codificacao)
        except (UnicodeDecodeError, LookupError):
            continue

    # Nenhuma serve por inteiro: fica a que menos estraga
    return bruto.decode("cp1252", errors="replace")


def preparar_xml(bruto):
    """Converte entidades HTML que o XML não reconhece (&nbsp;, &eacute;, ...).

    Os feeds noticiosos incluem com frequência entidades definidas em HTML mas não
    em XML, que fariam falhar a leitura de todo o feed.
    """
    texto = descodificar(bruto)
    reservadas = {"amp", "lt", "gt", "quot", "apos"}

    def trocar(m):
        nome = m.group(1)
        if nome in reservadas:
            return m.group(0)
        ponto = name2codepoint.get(nome)
        return chr(ponto) if ponto else " "

    return re.sub(r"&([a-zA-Z][a-zA-Z0-9]{1,31});", trocar, texto)


def endereco_do_item(item):
    """Endereço do artigo, seja qual for a forma que o feed usa.

    Em RSS o endereço vem no texto de <link>; em Atom vem no atributo href de
    <link>, e o texto fica vazio. Algumas publicações põem-no no <guid> ou num
    <origLink>. Ler apenas o texto de <link> deixava a coluna da ligação vazia
    nesses feeds — e no Excel o resumo, sendo longo, transbordava para o lugar
    dela, o que parecia uma repetição.
    """
    def limpo(v):
        v = (v or "").strip()
        return v if v.startswith("http") else ""

    # RSS: <link>https://…</link>
    achado = limpo(item.findtext("link"))
    if achado:
        return achado

    # Atom: <link rel="alternate" href="https://…"/>
    for no in item.iter():
        if not no.tag.endswith("link"):
            continue
        achado = limpo(no.get("href"))
        if achado and no.get("rel", "alternate") == "alternate":
            return achado

    # Alternativas frequentes
    for etiqueta in ("guid", "origLink", "id"):
        achado = limpo(item.findtext(etiqueta))
        if achado:
            return achado

    return ""


def extrair_itens(xml_bruto):
    """Devolve a lista de notícias de um feed RSS ou Atom."""
    raiz = ElementTree.fromstring(preparar_xml(xml_bruto))
    itens = []
    nos = list(raiz.iter("item"))
    if not nos:
        # Atom: os artigos vêm em <entry>, não em <item>
        nos = [n for n in raiz.iter() if n.tag.endswith("}entry") or n.tag == "entry"]
    for item in nos:
        titulo = limpar(item.findtext("title"))
        ligacao = endereco_do_item(item)

        # O <source> do item só é de fiar em feeds de agregador, onde cada
        # artigo vem de um órgão diferente. Nos feeds próprios de cada órgão —
        # que são todos os nossos — muitos publicadores usam esse campo para
        # outra coisa: a RFI mete lá o crédito da fotografia («© Reuters»,
        # «AFP - HENRY NICHOLLS») e o Expresso das Ilhas mete o título de outra
        # peça. Isso enchia a lista de publicações de nomes que não são
        # publicações. O nome do feed, que é o do órgão, é o que vale.
        fonte_no = item.find("source")
        fonte = limpar(fonte_no.text) if fonte_no is not None else ""
        dominio = ""
        if fonte_no is not None and fonte_no.get("url"):
            dominio = urlparse(fonte_no.get("url")).netloc

        # O Google acrescenta " - Fonte" ao fim do título; retira-se para não duplicar.
        if " - " in titulo:
            cabeca, cauda = titulo.rsplit(" - ", 1)
            if not fonte:
                titulo, fonte = cabeca, cauda
            elif cauda.strip().lower() == fonte.lower():
                titulo = cabeca

        data = None
        # As etiquetas de data variam com a plataforma. Além das três do costume,
        # muitos feeds portugueses (o Notícias ao Minuto, entre outros) datam os
        # artigos em <dc:date>, do Dublin Core — que o ElementTree devolve com o
        # espaço de nomes à cabeça. Sem esta linha, esses artigos ficavam SEM
        # DATA: eram lidos e marcados, entravam no retrato do dia, mas o arquivo,
        # que arruma por dia, descartava-os. Setecentos artigos por recolha a
        # desaparecerem em silêncio, e a publicação a parecer morta no painel.
        bruta = (item.findtext("pubDate") or item.findtext("published")
                 or item.findtext("updated") or item.findtext("date")
                 or item.findtext("{http://purl.org/dc/elements/1.1/}date"))
        if bruta:
            bruta = bruta.strip()
            try:
                data = para_lisboa(parsedate_to_datetime(bruta))
            except (TypeError, ValueError):
                # O Dublin Core usa ISO 8601 («2026-08-19T00:12:00Z»), que o
                # leitor de datas de correio não entende.
                try:
                    data = para_lisboa(datetime.fromisoformat(
                        bruta.replace("Z", "+00:00")))
                except ValueError:
                    data = None
            if data:
                # Algumas publicações datam artigos alguns minutos à frente do
                # relógio. Sem isto, o painel mostrava notícias com hora
                # posterior à atual, o que faz duvidar de tudo o resto.
                agora_ = agora_lisboa()
                if data > agora_:
                    data = agora_

        resumo = limpar(item.findtext("description") or item.findtext("summary")
                        or item.findtext("content"))
        # O Google devolve, na descrição, uma repetição do título seguida do nome da
        # publicação, ou uma lista de artigos relacionados. Nesses casos não há resumo.
        simplificar = lambda t: re.sub(r"[^a-z0-9áàâãéêíóôõúç ]", "", t.lower()).strip()  # noqa: E731
        if not resumo or simplificar(resumo).startswith(simplificar(titulo)[:45]):
            resumo = ""

        itens.append({"data": data, "fonte": fonte, "dominio": dominio,
                      "titulo": titulo, "resumo": resumo, "ligacao": ligacao,
                      "imagem": imagem_do_item(item)})
    return itens


def imagem_do_item(item):
    """Endereço da imagem que o feed associa ao artigo, se houver.

    As publicações declaram-na de várias formas — media:content, media:thumbnail,
    enclosure ou uma etiqueta img dentro da descrição. Percorrem-se todas, pela
    ordem em que costumam dar melhor resultado.
    """
    for etiqueta in ("content", "thumbnail"):
        for no in item.iter():
            if no.tag.endswith("}" + etiqueta) or no.tag == "media:" + etiqueta:
                url = no.get("url") or ""
                if url.startswith("http"):
                    return url

    fecho = item.find("enclosure")
    if fecho is not None:
        tipo = (fecho.get("type") or "")
        url = fecho.get("url") or ""
        if url.startswith("http") and (tipo.startswith("image") or not tipo):
            return url

    for campo in ("description", "{http://purl.org/rss/1.0/modules/content/}encoded"):
        bruto = item.findtext(campo) or ""
        achado = re.search(r'<img[^>]+src=["\']([^"\']+)["\']', bruto)
        if achado and achado.group(1).startswith("http"):
            return achado.group(1)

    return ""


def _sem_acentos(t):
    return unicodedata.normalize("NFD", t.lower()).encode("ascii", "ignore").decode()


def _formas(palavra, original):
    """Formas de superfície que a palavra pode assumir num título.

    Em vez de cortar a palavra e aceitar qualquer terminação — que fazia "IRS"
    casar com "irmão" e "pensões" com "pensou" — geram-se explicitamente o
    singular e o plural, e só se admite variação de sufixo em palavras longas.
    """
    # Siglas escrevem-se como estão: IRS, CPLP, AIMA, NATO, SIADAP
    if _e_sigla(original):
        return [re.escape(palavra)]

    formas = {palavra}

    # Cargos governativos admitem os dois géneros: uma expressão escrita como
    # "ministro da Saúde" casa também com "ministra da Saúde", e sobrevive
    # assim às remodelações sem se reescrever a lista. As formas são exatas
    # — géneros e plurais, sem o sufixo de derivação das palavras longas,
    # que faria "ministra" casar com "ministradas".
    _GENERO = {"ministro", "ministra", "secretario", "secretaria"}
    if palavra in _GENERO:
        raiz = palavra[:-1]
        return [re.escape(raiz + fim) for fim in ("o", "a", "os", "as")]
    if palavra.endswith("oes"):
        formas.add(palavra[:-3] + "ao")
    elif palavra.endswith("ao"):
        formas.add(palavra[:-2] + "oes")
    elif palavra.endswith("ais"):
        formas.add(palavra[:-3] + "al")
    elif palavra.endswith("al"):
        formas.add(palavra[:-2] + "ais")
    elif palavra.endswith("eis"):
        formas.add(palavra[:-3] + "el")
    elif palavra.endswith("el"):
        formas.add(palavra[:-2] + "eis")
    elif palavra.endswith("res"):
        formas.add(palavra[:-2])
    elif palavra.endswith("r"):
        formas.add(palavra + "es")
    elif palavra.endswith("s"):
        formas.add(palavra[:-1])
    else:
        formas.add(palavra + "s")

    # Palavras longas admitem variação de género e derivação curta:
    # escola → escolar, escolares; ferrovia → ferroviário
    sufixo = r"\w{0,3}" if len(palavra) >= 6 else ""
    return [re.escape(f) + sufixo for f in sorted(formas)]


def _e_sigla(palavra):
    return palavra.isupper() and len(palavra) <= 8


def contem_expressao(texto, expressao, texto_original=None):
    """Procura a expressão no texto, aceitando singular e plural.

    Sem isto, "medicamentos" não encontraria "medicamento" — e a maioria dos
    títulos usa o singular.

    Nas siglas exige-se maiúscula no texto original quando este é fornecido:
    "NATO" não deve casar com "líder nato", nem "IRS" com "irs".
    """
    palavras_orig = expressao.split()

    if texto_original:
        for sigla in (p for p in palavras_orig if _e_sigla(p)):
            if not re.search(r"(^|\W)" + re.escape(sigla) + r"(\W|$)", texto_original):
                return False

    palavras = [p for p in _sem_acentos(expressao).split() if p]
    if not palavras:
        return False

    partes = []
    for i, p in enumerate(palavras):
        original = palavras_orig[i] if i < len(palavras_orig) else p
        alternativas = _formas(p, original)
        partes.append("(?:" + "|".join(alternativas) + ")")

    padrao = r"\s+".join(partes)
    return re.search(r"(^|\W)" + padrao + r"(\W|$)", texto) is not None


# Domínios das publicações que escrevem em português. As restantes entram no
# corpus da pesquisa por termo, mas não são classificadas por área: as nossas
# palavras-chave são portuguesas e, aplicadas a outra língua, produzem
# coincidências falsas — "bolsas" apanhava "la bolsa española".
SUFIXOS_PORTUGUES = (".pt", ".ao", ".mz", ".cv", ".st", ".gw", ".tl", ".br")
DOMINIOS_PORTUGUES = (
    "noticiasaominuto.com", "cnnportugal.iol.pt", "eco.sapo.pt", "sapo.pt",
    "impresa.pt", "medialivre.pt", "lusa.pt", "novojornal.co.ao", "cartamz.com",
    "agenciabrasil.ebc.com.br", "folha.uol.com.br",
    # Publicação portuguesa em domínio .com — sem isto, o Ambiente Magazine era
    # lido mas nunca classificado, e o Ambiente e Energia perdia a única fonte
    # especializada que tem.
    "ambientemagazine.com",
    # Edições em português de publicações estrangeiras. A France 24 saiu desta
    # lista quando a fonte passou a ser a edição inglesa (a portuguesa não
    # existe): manter-la aqui mandava aplicar palavras-chave portuguesas a
    # texto inglês, que é exatamente o que produz coincidências falsas.
    "pt.euronews.com", "dw.com", "rfi.fr",
)


def escreve_em_portugues(dominio):
    """Verdadeiro se a publicação escreve em português."""
    d = (dominio or "").lower().replace("www.", "")
    if not d:
        return True                        # sem domínio conhecido, não se exclui
    if any(d.endswith(sufixo) for sufixo in SUFIXOS_PORTUGUES):
        return True
    return any(d == x or d.endswith("." + x) for x in DOMINIOS_PORTUGUES)


def marcar_por_areas(it, alvo):
    """Devolve as áreas e as palavras-chave que este artigo satisfaz.

    É a mesma regra que se aplicaria numa condição de um agregador de feeds:
    procura da expressão no título e no resumo. Um artigo pode pertencer a mais
    do que uma área.
    """
    bruto = it["titulo"] + " " + (it["resumo"] or "")
    texto = _sem_acentos(bruto)
    achados = []
    for ident, nome, grupo, palavras, excluir in alvo:
        if any(contem_expressao(texto, e, bruto) for e in excluir):
            continue
        casadas = [p for p in palavras if contem_expressao(texto, p, bruto)]
        if casadas:
            achados.append((nome, GRUPOS[grupo], casadas))
    return achados


def recolher_fontes(alvo, dias=7, pausa=0.4, internacionais=True, lusofonas=True):
    """Lê os feeds das publicações e marca os artigos pelas áreas governativas.

    Um artigo é recolhido por ser de uma fonte conhecida, e não por corresponder
    a uma pesquisa.
    """
    limite = agora_lisboa() - timedelta(days=dias)
    encontrados, lidos, falhas = {}, 0, []
    # Relatório por fonte de CADA recolha: quantos artigos deu, quantos marcou,
    # por que via e com que erro. Sem isto, uma publicação que deixa de
    # responder desaparece em silêncio — só se via o efeito, semanas depois, ao
    # dar pela falta dela nos quadros.
    relatorio = []
    # Domínios já lidos com êxito nesta recolha, para o recurso não duplicar
    # publicações que têm mais do que uma entrada na lista.
    dominios_lidos = set()
    sem_ligacao_por_fonte = {}
    # Todos os artigos lidos, marcados ou não: é este o corpus que a pesquisa
    # por termo livre percorre. Sem ele, procurar "Ceuta" só encontraria o que
    # já tivesse sido marcado por uma palavra-chave de área.
    todos = {}
    lista = list(FONTES)
    origens_lista = {n: "nacionais" for n, _, _ in FONTES}
    if lusofonas:
        lista += FONTES_LUSOFONAS
        origens_lista.update({n: "lusofonas" for n, _, _ in FONTES_LUSOFONAS})
    if internacionais:
        lista += FONTES_INTERNACIONAIS
        origens_lista.update({n: "internacionais" for n, _, _ in FONTES_INTERNACIONAIS})

    for i, (nome_fonte, dominio, url) in enumerate(lista, 1):
        via_google = nome_fonte in VIA_GOOGLE
        origem_da_lista = origens_lista.get(nome_fonte, "nacionais")
        print(f"[{i}/{len(lista)}] {nome_fonte}"
              f"{' (via Google Notícias)' if via_google else ''}…",
              end=" ", flush=True)
        recurso = ""
        try:
            itens = extrair_itens(ler_feed(url_via_google(dominio) if via_google else url))
        except Exception as erro:                              # noqa: BLE001
            itens, falha = None, str(erro)

        # RECURSO AUTOMÁTICO. Um feed direto pode falhar hoje e responder
        # amanhã: há órgãos que aceitam um pedido isolado e recusam quem os
        # visita de duas em duas horas. Quando a leitura direta falha ou vem
        # vazia, tenta-se de imediato a via Google para o mesmo domínio, em vez
        # de perder o dia inteiro dessa publicação.
        #
        # Só que o recurso é por DOMÍNIO, e há domínios com mais do que uma
        # entrada: as quatro secções do Público e a Lusa Internacional. Sendo
        # os feeds das secções endereços descontinuados, o recurso ia buscar ao
        # Google a mesma pesquisa site:publico.pt quatro vezes, e o jornal
        # passava a aparecer no painel como cinco publicações diferentes. Se o
        # domínio já foi lido com êxito nesta recolha, não se recorre: o que
        # essa pesquisa traria já lá está.
        if not via_google and not itens and dominio.lower() not in dominios_lidos:
            try:
                alternativos = extrair_itens(ler_feed(url_via_google(dominio)))
            except Exception:                                  # noqa: BLE001
                alternativos = []
            if alternativos:
                itens, recurso = alternativos, "google"

        if itens is None:
            print(f"falhou ({falha})")
            falhas.append((nome_fonte, falha))
            relatorio.append({"fonte": nome_fonte, "dominio": dominio,
                              "origem": origem_da_lista, "pt": escreve_em_portugues(dominio),
                              "via": "google" if via_google else "direta",
                              "lidos": 0, "marcados": 0, "erro": falha[:160]})
            continue
        if recurso:
            print("(direta sem resultado; via Google)", end=" ", flush=True)

        lidos += len(itens)
        marcados = 0
        sem_endereco = 0
        for it in itens:
            if it["data"] and it["data"] < limite:
                continue

            if not it["ligacao"]:
                sem_endereco += 1

            chave_geral = (it["titulo"] or "").lower()
            if chave_geral and chave_geral not in todos:
                todos[chave_geral] = {
                    "data": it["data"],
                    "fonte": nome_fonte or it["fonte"],
                    "dominio": it["dominio"] or dominio,
                    "titulo": it["titulo"],
                    "resumo": (it["resumo"] or "")[:240],
                    "ligacao": it["ligacao"],
                    "imagem": it.get("imagem", ""),
                }

            # Publicações noutra língua ficam pelo corpus da pesquisa: as
            # palavras-chave portuguesas, aplicadas a outra língua, dariam
            # coincidências falsas.
            if not escreve_em_portugues(it["dominio"] or dominio):
                continue
            for nome_area, grupo_nome, palavras in marcar_por_areas(it, alvo):
                chave = (nome_area, it["titulo"].lower())
                if chave in encontrados:
                    encontrados[chave]["palavras"].update(palavras)
                    continue
                encontrados[chave] = {
                    "area": nome_area, "grupo": grupo_nome, "data": it["data"],
                    "fonte": nome_fonte or it["fonte"], "dominio": it["dominio"] or dominio,
                    "titulo": it["titulo"], "resumo": it["resumo"],
                    "ligacao": it["ligacao"], "imagem": it.get("imagem", ""),
                    "palavras": set(palavras),
                }
                marcados += 1
        # As publicações que não indicam o endereço no feed ficam identificadas:
        # sem isto, só se via o efeito — notícias sem ligação — e não a causa.
        aviso = f", {sem_endereco} SEM ENDEREÇO" if sem_endereco else ""
        print(f"{len(itens)} artigos, {marcados} marcados{aviso}")
        if itens:
            dominios_lidos.add(dominio.lower())
        relatorio.append({"fonte": nome_fonte, "dominio": dominio,
                          "origem": origem_da_lista, "pt": escreve_em_portugues(dominio),
                          "via": "google" if via_google else
                                 ("google (recurso)" if recurso else "direta"),
                          "lidos": len(itens), "marcados": marcados, "erro": ""})
        if sem_endereco:
            sem_ligacao_por_fonte[nome_fonte] = sem_endereco
        if i < len(lista):
            time.sleep(pausa)

    print(f"\n{lidos} artigos lidos · {len(todos)} distintos · "
          f"{len(encontrados)} marcados por área")

    # As ligações das publicações lidas via Google Notícias são
    # reencaminhamentos; resolvem-se aqui para o endereço do jornal, com o
    # resolvedor da era Google. Só se tenta uma vez por ligação, em paralelo,
    # e o que não resolver fica com o reencaminhamento — que também abre.
    pendentes = sorted({d["ligacao"]
                        for d in list(encontrados.values()) + list(todos.values())
                        if d["ligacao"] and "news.google." in d["ligacao"]})
    if pendentes:
        print(f"A resolver {len(pendentes)} ligações do Google Notícias…",
              end=" ", flush=True)
        with ThreadPoolExecutor(max_workers=8) as piscina:
            mapa = dict(zip(pendentes, piscina.map(endereco_do_artigo, pendentes)))
        trocadas = 0
        for d in list(encontrados.values()) + list(todos.values()):
            novo = mapa.get(d["ligacao"])
            if novo and novo != d["ligacao"]:
                d["ligacao"] = novo
                trocadas += 1
        print(f"{trocadas} convertidas em endereço do jornal")

    if sem_ligacao_por_fonte:
        total = sum(sem_ligacao_por_fonte.values())
        print(f"\n{total} artigos sem endereço no feed, nestas publicações:")
        for nome, quantos in sorted(sem_ligacao_por_fonte.items(), key=lambda x: -x[1]):
            print(f"   {nome}: {quantos}")
        print("Estas notícias entram na mesma, mas sem ligação para o artigo.")
    # O relatório por fonte fica em ficheiro, não só no registo da execução:
    # é a diferença entre saber e ter de ir procurar.
    try:
        mudas = [r for r in relatorio if not r["lidos"]]
        with open("fontes-recolha.json", "w", encoding="utf-8") as destino:
            json.dump({"recolhido": agora_lisboa().strftime("%Y-%m-%d %H:%M"),
                       "fontes": len(relatorio),
                       "sem_artigos": len(mudas),
                       "por_recurso": len([r for r in relatorio
                                           if r["via"] == "google (recurso)"]),
                       "detalhe": sorted(relatorio,
                                         key=lambda r: (r["lidos"], r["fonte"]))},
                      destino, ensure_ascii=False, indent=1)
        if mudas:
            print("Sem um único artigo nesta recolha: "
                  + ", ".join(r["fonte"] for r in mudas))
    except OSError:
        pass

    return list(encontrados.values()), falhas, list(todos.values())


def recolher_palavras(periodo, alvo, pausa=1.0, teto=60):
    """Uma consulta por palavra-chave, além da consulta da área.

    É isto que permite ao painel responder a pesquisas por palavra-chave sem
    interrogar o serviço: os resultados de cada termo ficam nos ficheiros.
    Cada notícia guarda as palavras-chave que a trouxeram.
    """
    encontradas = {}
    total = sum(len(a[3]) for a in alvo)
    feitas = 0

    for ident, nome, grupo, palavras, excluir in alvo:
        for palavra in palavras:
            feitas += 1
            q = f'"{palavra}"'
            if periodo:
                q += f" when:{periodo}"
            try:
                itens = extrair_itens(ler_feed(url_feed(q)))[:teto]
            except Exception as erro:                          # noqa: BLE001
                print(f"  [{feitas}/{total}] {palavra}: falhou ({erro})")
                continue
            novas = 0
            for it in itens:
                # o serviço faz correspondência aproximada: confirma-se o termo
                bruto = it["titulo"] + " " + (it["resumo"] or "")
                if not contem_expressao(_sem_acentos(bruto), palavra, bruto):
                    continue
                chave = (nome, it["titulo"].lower(), (it["fonte"] or "").lower())
                if chave in encontradas:
                    encontradas[chave]["palavras"].add(palavra)
                else:
                    registo = dict(it)
                    registo.update({"area": nome, "grupo": GRUPOS[grupo], "palavras": {palavra}})
                    encontradas[chave] = registo
                    novas += 1
            print(f"  [{feitas}/{total}] {palavra}: {novas} novas")
            time.sleep(pausa)

    return list(encontradas.values())


def recolher(periodo, apenas=None, so_nacionais=True, pausa=1.5):
    linhas, falhas = [], []
    vistos = set()
    alvo = [a for a in AREAS if apenas is None or a[0] == apenas]

    for i, (ident, nome, grupo, palavras, excluir) in enumerate(alvo, 1):
        q = consulta(palavras, excluir, periodo, so_nacionais)
        print(f"[{i}/{len(alvo)}] {nome}…", end=" ", flush=True)
        try:
            itens = extrair_itens(ler_feed(url_feed(q)))
        except Exception as erro:                              # noqa: BLE001
            print(f"falhou ({erro})")
            falhas.append((nome, str(erro)))
            continue

        novos = 0
        for it in itens:
            chave = (ident, it["ligacao"])
            if chave in vistos:
                continue
            vistos.add(chave)
            linhas.append([nome, GRUPOS[grupo], it["data"], it["fonte"], it["dominio"],
                           it["titulo"], it["ligacao"]])
            novos += 1
        print(f"{novos} notícias")
        if i < len(alvo):
            time.sleep(pausa)

    # Mais recentes primeiro; as notícias sem data reconhecível ficam no fim.
    com_data = [l for l in linhas if l[2] is not None]
    sem_data = [l for l in linhas if l[2] is None]
    com_data.sort(key=lambda l: l[2], reverse=True)
    return com_data + sem_data, falhas


# ---------------------------------------------------------------------------
# Escrita do Excel
# ---------------------------------------------------------------------------
def gravar(linhas, falhas, caminho, periodo, so_nacionais):
    fonte = lambda **k: Font(name="Arial", **k)                # noqa: E731
    cab = PatternFill("solid", fgColor=AZUL)
    alt = PatternFill("solid", fgColor=CINZA)
    borda = Border(bottom=Side(style="thin", color="D5DCE4"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Notícias"
    cabecalhos = ["Área governativa", "Agrupamento temático", "Data de publicação", "Fonte",
                  "Domínio", "Título", "Resumo", "Ligação"]
    for j, h in enumerate(cabecalhos, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = fonte(bold=True, color="FFFFFF", size=10)
        c.fill = cab
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    # A folha tem as colunas do cabeçalho e mais nenhuma: as linhas trazem também
    # as palavras-chave e a imagem, que servem ao arquivo mas não à folha.
    for i, l in enumerate(linhas, 2):
        for j, v in enumerate(l[:len(cabecalhos)], 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = fonte(size=10)
            c.border = borda
            c.alignment = Alignment(vertical="top", wrap_text=(j in (6, 7)))
            if i % 2 == 0:
                c.fill = alt
            if j == 3 and v is not None:
                c.number_format = "yyyy-mm-dd hh:mm"
            if j == 8 and v:
                c.hyperlink = v
                c.font = fonte(size=10, color="2B5683", underline="single")

    for j, w in enumerate([30, 28, 19, 24, 22, 54, 54, 46], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{max(2, len(linhas) + 1)}"

    # --- Folha de síntese ---
    wr = wb.create_sheet("Síntese")
    wr.column_dimensions["A"].width = 44
    wr.column_dimensions["B"].width = 16
    wr["A1"] = "Síntese da recolha"
    wr["A1"].font = fonte(size=13, bold=True, color=AZUL)

    contexto = [
        ("Data da recolha", agora_lisboa().strftime("%Y-%m-%d %H:%M")),
        ("Janela temporal", periodo or "sem limite"),
        ("Restrição a fontes nacionais", "sim" if so_nacionais else "não"),
        ("Notícias recolhidas", len(linhas)),
        ("Áreas com falha na recolha", len(falhas)),
    ]
    for i, (t, v) in enumerate(contexto, start=3):
        wr.cell(row=i, column=1, value=t).font = fonte(size=10)
        wr.cell(row=i, column=2, value=v).font = fonte(size=10, bold=True)

    wr.cell(row=9, column=1, value="Notícias por área governativa").font = fonte(size=11, bold=True, color=AZUL)
    ultima = max(2, len(linhas) + 1)
    for i, (_, nome, _, _, _) in enumerate(AREAS, start=10):
        wr.cell(row=i, column=1, value=nome).font = fonte(size=10)
        c = wr.cell(row=i, column=2, value=f'=COUNTIF(Notícias!$A$2:$A${ultima},A{i})')
        c.font = fonte(size=10, bold=True)
        c.alignment = Alignment(horizontal="center")

    aviso = wr.cell(row=10 + len(AREAS) + 1, column=1,
                    value=("Nota: o serviço devolve no máximo cerca de 100 artigos por consulta. "
                           "As contagens medem exposição mediática indexada e não são comparáveis "
                           "entre áreas. A coluna Resumo fica vazia quando o feed não fornece texto "
                           "próprio, o que sucede na maioria dos casos."))
    aviso.font = fonte(size=9, italic=True, color="5B6068")
    aviso.alignment = Alignment(wrap_text=True, vertical="top")
    wr.merge_cells(start_row=aviso.row, start_column=1, end_row=aviso.row, end_column=2)
    wr.row_dimensions[aviso.row].height = 60

    if falhas:
        wf = wb.create_sheet("Falhas")
        wf.column_dimensions["A"].width = 34
        wf.column_dimensions["B"].width = 80
        wf["A1"], wf["B1"] = "Área", "Erro"
        for c in ("A1", "B1"):
            wf[c].font = fonte(bold=True, color="FFFFFF", size=10)
            wf[c].fill = cab
        for i, (nome, erro) in enumerate(falhas, 2):
            wf.cell(row=i, column=1, value=nome).font = fonte(size=10)
            wf.cell(row=i, column=2, value=erro).font = fonte(size=10)

    wb.save(caminho)


# ---------------------------------------------------------------------------
def principal():
    ap = argparse.ArgumentParser(description="Extrai as notícias do radar para Excel.")
    ap.add_argument("--periodo", default="7d",
                    help="janela temporal: 24h, 48h, 72h, 7d, 30d, 365d ou vazio")
    ap.add_argument("--area", default=None, help="identificador de uma única área (ex.: saude)")
    ap.add_argument("--todas-as-fontes", action="store_true",
                    help="não excluir os domínios estrangeiros")
    ap.add_argument("--saida", default=None, help="nome do ficheiro Excel a criar")
    ap.add_argument("--json", default=None,
                    help="grava também um ficheiro JSON com as notícias (para publicar junto ao painel)")
    ap.add_argument("--sem-resolver-ligacoes", action="store_true",
                    help="não converter os reencaminhamentos do Google no endereço do jornal")
    ap.add_argument("--historico", default=None,
                    help="ficheiro JSON de série diária, atualizado a cada recolha")
    ap.add_argument("--arquivo", default=None,
                    help="ficheiro JSON com as notícias dos últimos dias, acumulado a cada recolha")
    ap.add_argument("--corpus", default=None,
                    help="ficheiro com todos os artigos lidos, para a pesquisa livre")
    ap.add_argument("--mensal", default=None,
                    help="pasta do arquivo permanente, um ficheiro comprimido por mês")
    ap.add_argument("--dias-arquivo", type=int, default=7,
                    help="dias a manter no arquivo (predefinição: 7)")
    ap.add_argument("--por-palavra", action="store_true",
                    help="recolher também uma consulta por cada palavra-chave")
    ap.add_argument("--fontes", action="store_true",
                    help="ler os feeds das publicações e marcar por área (método de agregador)")
    ap.add_argument("--sem-internacionais", action="store_true",
                    help="ler apenas as publicações portuguesas, sem imprensa estrangeira")
    args = ap.parse_args()

    if args.area and args.area not in {a[0] for a in AREAS}:
        sys.exit(f"Área desconhecida: {args.area}")

    saida = args.saida or f"radar_noticias_{agora_lisboa():%Y%m%d_%H%M}.xlsx"
    so_nacionais = not args.todas_as_fontes

    alvo = [a for a in AREAS if args.area is None or a[0] == args.area]
    das_fontes = []

    if args.fontes:
        # Origem principal: os feeds das próprias publicações
        print("Leitura dos feeds das publicações:")
        das_fontes, falhas, todos_lidos = recolher_fontes(
            alvo, args.dias_arquivo,
            internacionais=not args.sem_internacionais,
            lusofonas=not args.sem_internacionais)
        campos = ["area", "grupo", "data", "fonte", "dominio", "titulo", "resumo", "ligacao"]
        # As palavras-chave e a imagem seguem com a linha: sem elas, o arquivo
        # ficava sem imagem nas notícias marcadas por área, e só as mostrava nos
        # resultados da pesquisa por termo.
        linhas = [[n["area"], n["grupo"], n["data"], n["fonte"], n["dominio"],
                   n["titulo"], n["resumo"], n["ligacao"],
                   sorted(n.get("palavras") or []), n.get("imagem", "")]
                  for n in das_fontes]
        linhas.sort(key=lambda l: (l[2] is None, l[2]), reverse=True)
    else:
        todos_lidos = []
        linhas, falhas = recolher(args.periodo, args.area, so_nacionais)
        if not args.sem_resolver_ligacoes and linhas:
            linhas = resolver_ligacoes(linhas)
        linhas = [l[:6] + [""] + l[6:] for l in linhas]   # coluna de resumo vazia
    gravar(linhas, falhas, saida, args.periodo, so_nacionais)
    print(f"\n{len(linhas)} notícias gravadas em {saida}")

    vistos_anteriores = ligacoes_da_recolha_anterior(args.json) if (args.json and args.historico) else {}

    if args.json:
        campos = ["area", "grupo", "data", "fonte", "dominio", "titulo", "resumo", "ligacao"]
        noticias = []
        for l in linhas:
            registo = dict(zip(campos, l))
            registo["data"] = registo["data"].strftime("%Y-%m-%d %H:%M") if registo["data"] else ""
            noticias.append(registo)
        MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
                 "agosto", "setembro", "outubro", "novembro", "dezembro"]
        agora = agora_lisboa()
        if args.fontes:
            for registo, n in zip(noticias, das_fontes):
                registo["palavras"] = sorted(n["palavras"])
        pacote = {
            "gerado": f"{agora.day} de {MESES[agora.month - 1]} de {agora.year}, {agora:%Hh%M}",
            "origem": "feeds das publicações" if args.fontes else "Google Notícias",
            "periodo": args.periodo or "sem limite",
            "noticias": noticias,
        }
        with open(args.json, "w", encoding="utf-8") as destino:
            json.dump(pacote, destino, ensure_ascii=False, indent=1)
        print(f"{len(noticias)} notícias gravadas em {args.json}")

    por_palavra = list(das_fontes)

    if args.por_palavra:
        print("\nRecolha por palavra-chave:")
        por_palavra = recolher_palavras(args.periodo, alvo)
        if not args.sem_resolver_ligacoes and por_palavra:
            enderecos = {n["ligacao"] for n in por_palavra}
            print(f"A resolver {len(enderecos)} ligações…", end=" ", flush=True)
            with ThreadPoolExecutor(max_workers=8) as piscina:
                mapa = dict(zip(enderecos, piscina.map(endereco_do_artigo, enderecos)))
            for n in por_palavra:
                n["ligacao"] = mapa.get(n["ligacao"], n["ligacao"])
            print("feito")
        print(f"{len(por_palavra)} notícias com palavra-chave identificada")

    if args.arquivo:
        atualizar_arquivo(args.arquivo, linhas, args.dias_arquivo, por_palavra)

    if args.corpus and todos_lidos:
        gravar_corpus(args.corpus, todos_lidos, args.dias_arquivo)

    if args.mensal:
        guardar_mensal(args.mensal, linhas)
        if todos_lidos:
            guardar_mensal_integral(args.mensal, todos_lidos, linhas)

    if args.historico:
        atualizar_historico(args.historico, linhas, args.periodo, vistos_anteriores,
                            arquivo=args.arquivo)
    if falhas:
        print(f"{len(falhas)} áreas falharam — ver a folha Falhas.")


def ligacoes_da_recolha_anterior(caminho_json):
    """Ligações da recolha anterior, para distinguir o que é novo.

    As janelas de 24 horas não são perfeitamente disjuntas: o operador de tempo
    do Google atua sobre a data de indexação e um artigo reindexado volta a
    aparecer. Comparar com a véspera é o que permite contar notícias novas em
    vez de contar repetições.
    """
    if not os.path.exists(caminho_json):
        return {}
    try:
        with open(caminho_json, encoding="utf-8") as origem:
            anterior = json.load(origem)
    except (json.JSONDecodeError, OSError):
        return {}
    vistos = {}
    for n in anterior.get("noticias", []):
        vistos.setdefault(n.get("area", ""), set()).add(n.get("ligacao", ""))
    return vistos


# Caractere que substitui os bytes que não puderam ser descodificados. Aparecia
# no lugar dos acentos nos feeds que não publicam em UTF-8, antes de a leitura
# passar a respeitar a codificação declarada.
SUBSTITUTO = "\ufffd"


def texto_estragado(registo):
    """Verdadeiro se o registo está inutilizável e deve voltar a ser recolhido.

    Dois casos. O primeiro são os caracteres que não puderam ser lidos, de
    quando a leitura não respeitava a codificação de cada feed. O segundo são as
    ligações que não são endereços: durante algum tempo, um desalinhamento de
    campos gravou o resumo no lugar da ligação.

    Descartá-los é o que permite substituí-los, porque a desduplicação mantém o
    registo mais antigo — sem isto, o defeituoso continuaria a impedir a entrada
    do correto.
    """
    if any(SUBSTITUTO in (registo.get(campo) or "")
           for campo in ("titulo", "resumo", "fonte")):
        return True

    ligacao = (registo.get("ligacao") or "").strip()
    return bool(ligacao) and not ligacao.startswith("http")


def atualizar_arquivo(caminho, linhas, dias=7, por_palavra=None):
    """Acumula as notícias dos últimos dias num único ficheiro.

    É o que permite ao painel responder a pesquisas por palavra-chave e a
    janelas de vários dias sem depender de serviços externos: em vez de
    interrogar o serviço, filtra este arquivo.
    """
    # A ordem tem de ser a das linhas produzidas pela recolha. Faltava aqui o
    # resumo, pelo que a sétima posição — que é o resumo — ia parar ao campo da
    # ligação: no Excel, a coluna da ligação aparecia com o texto do resumo.
    campos = ["area", "grupo", "data", "fonte", "dominio", "titulo", "resumo",
              "ligacao", "palavras", "imagem"]
    limite = (agora_lisboa() - timedelta(days=dias)).strftime("%Y-%m-%d")

    anteriores = []
    if os.path.exists(caminho):
        try:
            with open(caminho, encoding="utf-8") as origem:
                anteriores = json.load(origem).get("noticias", [])
        except (json.JSONDecodeError, OSError):
            anteriores = []

    # Registos guardados antes de a leitura respeitar a codificação de cada feed
    # trazem losangos no lugar dos acentos. Saem daqui: a recolha seguinte volta
    # a trazê-los legíveis, e mantê-los duplicaria a mesma notícia.
    estragados = sum(1 for r in anteriores if texto_estragado(r))
    if estragados:
        anteriores = [r for r in anteriores if not texto_estragado(r)]
        print(f"arquivo: {estragados} registos com caracteres ilegíveis descartados")

    def registo(l):
        r = dict(zip(campos, l))
        r["data"] = r["data"].strftime("%Y-%m-%d %H:%M") if r["data"] else ""
        r["palavras"] = sorted(r["palavras"]) if r.get("palavras") else []
        if not r.get("resumo"):
            r.pop("resumo", None)
        if not r.get("imagem"):
            r.pop("imagem", None)
        return r

    novas = [registo(l) for l in linhas]
    for n in (por_palavra or []):
        registo_novo = {
            "area": n["area"], "grupo": n["grupo"],
            "data": n["data"].strftime("%Y-%m-%d %H:%M") if n["data"] else "",
            "fonte": n["fonte"], "dominio": n["dominio"], "titulo": n["titulo"],
            "ligacao": n["ligacao"], "palavras": sorted(n["palavras"]),
        }
        if n.get("resumo"):
            registo_novo["resumo"] = n["resumo"]
        if n.get("imagem"):
            registo_novo["imagem"] = n["imagem"]
        novas.append(registo_novo)

    juntas = anteriores + novas
    vistos, mantidas = {}, []
    for n in juntas:
        # Sem data reconhecida não entra: seria impossível saber se está na janela
        if not n.get("data") or n["data"][:10] < limite:
            continue
        chave = (n.get("area", ""), n.get("titulo", "").lower(), n.get("fonte", "").lower())
        if chave in vistos:
            # Já existe: em vez de descartar a repetição, aproveita-se dela o que
            # faltar ao registo guardado. É o que permite acrescentar imagem,
            # resumo ou ligação a notícias recolhidas antes de esses campos
            # passarem a ser gravados — sem isso, o registo antigo, incompleto,
            # continuaria a impedir a entrada do novo.
            anterior = vistos[chave]
            if n.get("palavras"):
                anterior["palavras"] = sorted(set(anterior.get("palavras", [])) | set(n["palavras"]))
            for campo in ("resumo", "imagem", "ligacao", "dominio", "grupo"):
                if n.get(campo) and not anterior.get(campo):
                    anterior[campo] = n[campo]
            continue
        vistos[chave] = n
        mantidas.append(n)

    mantidas.sort(key=lambda n: n.get("data", ""), reverse=True)
    with open(caminho, "w", encoding="utf-8") as destino:
        json.dump({"dias": dias, "atualizado": agora_lisboa().strftime("%Y-%m-%d %H:%M"),
                   "noticias": mantidas}, destino, ensure_ascii=False)
    print(f"arquivo de {dias} dias: {len(mantidas)} notícias em {caminho}")


# Publicações portuguesas cujo domínio não termina em .pt
DOMINIOS_PT = ("noticiasaominuto.com", "theportugalnews.com", "portugalresident.com",
               "agencialusa.com", "cnnportugal.iol.pt", "iol.pt", "eco.sapo.pt",
               "sapo.pt", "aeiou.pt", "impresa.pt", "medialivre.pt", "lusa.pt")
DOMINIOS_LUSOFONOS = (".ao", ".mz", ".cv", ".st", ".gw", ".tl", ".br")


def origem_da_fonte(dominio):
    """Portugal, lusofonia ou internacional, a partir do domínio.

    As mesmas três origens do painel e do relatório: sem isto, a série diária
    contaria de maneira diferente da consulta.
    """
    d = (dominio or "").lower().replace("www.", "")
    if not d:
        return "nacionais"
    if d.endswith(".pt") or any(d == x or d.endswith("." + x) for x in DOMINIOS_PT):
        return "nacionais"
    if any(d.endswith(t) for t in DOMINIOS_LUSOFONOS):
        return "lusofonas"
    return "internacionais"


def guardar_mensal(pasta, linhas):
    """Arquivo permanente, um ficheiro comprimido por mês.

    O arquivo corrente guarda sete dias e é o que o painel consulta; este é o
    depósito de longo prazo, que nada lê no dia a dia. Serve para, daqui a um
    ano, se poder voltar atrás e produzir estatísticas sobre o que foi noticiado.

    Cada linha é uma notícia em JSON — formato que se lê de forma incremental,
    sem carregar o ficheiro todo em memória. Comprimido, ocupa cerca de um
    décimo do que ocuparia em texto simples.
    """
    import gzip

    os.makedirs(pasta, exist_ok=True)
    por_mes = {}
    for l in linhas:
        data = l[2]
        if not data:
            continue
        mes = data.strftime("%Y-%m")
        por_mes.setdefault(mes, []).append({
            "area": l[0], "grupo": l[1],
            "data": data.strftime("%Y-%m-%d %H:%M"),
            "fonte": l[3], "dominio": l[4], "titulo": l[5],
            "resumo": l[6] or "", "ligacao": l[7],
            "palavras": sorted(l[8]) if len(l) > 8 and l[8] else [],
            # A imagem viaja com a notícia desde a recolha: guardá-la aqui é o
            # que permite mostrar as notícias antigas com o mesmo aspeto das
            # recentes, quando a janela do painel se alargar para além dos
            # sete dias do arquivo. Custa poucos bytes — é só o endereço.
            "imagem": (l[9] if len(l) > 9 else "") or "",
        })

    total = 0
    for mes, registos in sorted(por_mes.items()):
        caminho = os.path.join(pasta, f"{mes}.jsonl.gz")

        # Sem repetições: uma notícia recolhida em dois dias entra uma só vez
        vistos = set()
        anteriores = []
        if os.path.exists(caminho):
            with gzip.open(caminho, "rt", encoding="utf-8") as origem:
                for linha in origem:
                    try:
                        r = json.loads(linha)
                    except json.JSONDecodeError:
                        continue
                    if texto_estragado(r):
                        continue          # volta a entrar legível na recolha seguinte
                    anteriores.append(r)
                    vistos.add((r.get("area", ""), (r.get("titulo") or "").lower()))

        novas = [r for r in registos
                 if (r["area"], (r["titulo"] or "").lower()) not in vistos]
        if not novas:
            print(f"arquivo mensal {mes}: sem notícias novas")
            continue

        juntas = anteriores + novas
        juntas.sort(key=lambda r: r["data"], reverse=True)
        with gzip.open(caminho, "wt", encoding="utf-8") as destino:
            for r in juntas:
                destino.write(json.dumps(r, ensure_ascii=False) + "\n")

        tamanho = os.path.getsize(caminho) / 1024
        total += len(novas)
        print(f"arquivo mensal {mes}: +{len(novas)} novas, "
              f"{len(juntas)} ao todo ({tamanho:.0f} KB)")
    return total


def gravar_corpus(caminho, lidos, dias=7):
    """Todos os artigos lidos dos feeds, marcados ou não, dos últimos dias.

    É o que permite a pesquisa por termo livre encontrar matéria que nenhuma
    palavra-chave de área apanhou — o caso de Ceuta, que era noticiado sem que
    as expressões da Presidência o cobrissem.

    Guarda-se sem área e sem palavras-chave: é imprensa em bruto, não corpus
    classificado. O painel só o carrega quando alguém pesquisa por termo.
    """
    limite = (agora_lisboa() - timedelta(days=dias)).strftime("%Y-%m-%d")

    anteriores = []
    if os.path.exists(caminho):
        try:
            with open(caminho, encoding="utf-8") as origem:
                anteriores = json.load(origem).get("noticias", [])
        except (json.JSONDecodeError, OSError):
            anteriores = []

    estragados = sum(1 for r in anteriores if texto_estragado(r))
    if estragados:
        anteriores = [r for r in anteriores if not texto_estragado(r)]
        print(f"corpus: {estragados} registos com caracteres ilegíveis descartados")

    novos = []
    for n in lidos:
        novos.append({
            "data": n["data"].strftime("%Y-%m-%d %H:%M") if n["data"] else "",
            "fonte": n["fonte"], "dominio": n["dominio"],
            "titulo": n["titulo"], "resumo": n["resumo"],
            "ligacao": n["ligacao"], "imagem": n.get("imagem", ""),
        })

    vistos, mantidos = {}, []
    for n in novos + anteriores:
        if not n.get("data") or n["data"][:10] < limite:
            continue
        chave = (n.get("titulo") or "").lower()
        if chave in vistos:
            anterior = vistos[chave]
            for campo in ("resumo", "imagem", "ligacao", "dominio"):
                if n.get(campo) and not anterior.get(campo):
                    anterior[campo] = n[campo]
            continue
        vistos[chave] = n
        mantidos.append(n)

    mantidos.sort(key=lambda n: n["data"], reverse=True)
    with open(caminho, "w", encoding="utf-8") as destino:
        json.dump({"dias": dias, "atualizado": agora_lisboa().strftime("%Y-%m-%d %H:%M"),
                   "noticias": mantidos}, destino, ensure_ascii=False, indent=1)

    tamanho = os.path.getsize(caminho) / 1024 / 1024
    print(f"corpus de {dias} dias: {len(mantidos)} artigos em {caminho} ({tamanho:.1f} MB)")


def guardar_mensal_integral(pasta, lidos, linhas):
    """Depósito mensal também para as notícias NÃO marcadas (área vazia).

    O depósito guardava apenas o que as regras do dia marcaram. Mas as regras
    mudam — criou-se a área do Primeiro-Ministro, acrescentaram-se os cargos —
    e o que não foi marcado hoje pode ser exatamente o que uma regra futura
    procurará. Guardando tudo, os passos retroativos deixam de estar presos à
    janela de sete dias do corpus: passam a ter meses de profundidade.

    Uma notícia entra com área vazia apenas se nenhuma área a marcou: as
    marcadas já lá estão com as suas áreas, e o texto é o mesmo. Custa cerca
    de dois a três megabytes comprimidos por mês. A área vazia é ignorada por
    tudo o que conta por área (alertas, séries) e mantida pela revalidação.
    """
    import gzip

    os.makedirs(pasta, exist_ok=True)
    marcados = {(l[5] or "").lower() for l in linhas}

    por_mes = {}
    for n in lidos:
        data = n.get("data")
        titulo = (n.get("titulo") or "").strip()
        if not data or not titulo or titulo.lower() in marcados:
            continue
        por_mes.setdefault(data.strftime("%Y-%m"), []).append({
            "area": "", "grupo": "",
            "data": data.strftime("%Y-%m-%d %H:%M"),
            "fonte": n.get("fonte", ""), "dominio": n.get("dominio", ""),
            "titulo": titulo, "resumo": n.get("resumo") or "",
            "ligacao": n.get("ligacao", ""), "imagem": n.get("imagem") or "",
            "palavras": [],
        })

    total = 0
    for mes, registos in sorted(por_mes.items()):
        caminho = os.path.join(pasta, f"{mes}.jsonl.gz")
        vistos = set()
        if os.path.exists(caminho):
            with gzip.open(caminho, "rt", encoding="utf-8") as origem:
                for linha in origem:
                    try:
                        r = json.loads(linha)
                    except json.JSONDecodeError:
                        continue
                    vistos.add((r.get("titulo") or "").lower())
        novas = [r for r in registos if r["titulo"].lower() not in vistos]
        if not novas:
            continue
        with gzip.open(caminho, "at", encoding="utf-8") as destino:
            for r in novas:
                destino.write(json.dumps(r, ensure_ascii=False) + "\n")
        total += len(novas)
    if total:
        print(f"depósito mensal integral: +{total} notícias não marcadas")


def historico_do_arquivo(caminho, arquivo):
    """Reconstrói a série dos últimos dias a partir do arquivo acumulado.

    ANTES CONTAVA-SE MAL. Cada recolha registava o dia com base no que ela
    própria tinha lido dos feeds naquele instante — e um feed só expõe os seus
    últimos artigos. Como cada execução substituía o registo do dia, o valor que
    ficava era o da última recolha, e não a soma do que o dia trouxe. Daí a série
    mostrar cerca de um terço do que o arquivo tem.

    O arquivo, esse, acumula todas as recolhas e está desduplicado. Reconstruir a
    série a partir dele torna os dois números coerentes por construção: o que o
    painel de evolução soma passa a ser o mesmo que o radar mostra.

    Reescrevem-se todos os dias que o arquivo cobre — sete —, o que também
    corrige o dia anterior com o que foi publicado depois da última recolha.
    """
    try:
        with open(arquivo, encoding="utf-8") as origem:
            noticias = json.load(origem).get("noticias", [])
    except (json.JSONDecodeError, OSError) as erro:
        print(f"Série diária: não foi possível ler {arquivo} ({erro}).")
        return None

    por_dia = {}
    # Notícias DISTINTAS por dia: a mesma notícia deixa uma linha por área, e
    # somar as áreas dá marcações, não notícias. O painel mostra estas.
    distintas = {}
    for n in noticias:
        dia = (n.get("data") or "")[:10]
        area = n.get("area")
        if not dia or not area:
            continue
        lig = n.get("ligacao") or ""
        if lig:
            distintas.setdefault(dia, {})[lig] = origem_da_fonte(n.get("dominio"))
        registo = por_dia.setdefault(dia, {}).setdefault(area, {
            "noticias": 0, "novas": 0, "fontes": set(),
            "origens": {"nacionais": 0, "lusofonas": 0, "internacionais": 0},
            "palavras": {},
        })
        registo["noticias"] += 1
        registo["novas"] += 1
        if n.get("fonte"):
            registo["fontes"].add(n["fonte"])
        registo["origens"][origem_da_fonte(n.get("dominio"))] += 1
        for pal in (n.get("palavras") or []):
            registo["palavras"][pal] = registo["palavras"].get(pal, 0) + 1

    saida = {}
    for dia, areas in por_dia.items():
        contagem = {"total": 0, "nacionais": 0, "lusofonas": 0, "internacionais": 0}
        for o in distintas.get(dia, {}).values():
            contagem["total"] += 1
            contagem[o] += 1
        saida[dia] = {
            "distintas": contagem,
            "areas": {nome: {
                "noticias": v["noticias"], "novas": v["novas"], "fontes": len(v["fontes"]),
                "origens": v["origens"],
                "palavras": dict(sorted(v["palavras"].items(), key=lambda x: -x[1])),
            } for nome, v in sorted(areas.items())},
        }
    return saida


def atualizar_historico(caminho, linhas, periodo, vistos=None, arquivo=None):
    """Acrescenta à série diária o retrato de hoje, por área governativa.

    Guarda apenas agregados — notícias recolhidas, notícias publicadas nesse dia
    e publicações distintas. É deliberadamente pequeno: cresce cerca de 2 KB por
    dia e pode ser lido de uma vez pelo painel.

    A contagem do que é próprio do dia usa a DATA DE PUBLICAÇÃO, e não a
    comparação com a recolha anterior. Com recolhas de duas em duas horas, a
    segunda recolha do dia quase nada traria de novo face à primeira, e a série
    passava a mostrar zeros — foi o que sucedeu.
    """
    vistos = vistos or {}
    hoje = agora_lisboa().strftime("%Y-%m-%d")

    serie = {"atualizado": hoje, "dias": []}
    if os.path.exists(caminho):
        try:
            with open(caminho, encoding="utf-8") as origem:
                anterior = json.load(origem)
            if isinstance(anterior.get("dias"), list):
                serie["dias"] = [d for d in anterior["dias"] if d.get("data") != hoje]
        except (json.JSONDecodeError, OSError):
            pass                                   # série ilegível: recomeça-se

    # A série é reconstruída a partir do arquivo acumulado, que é o mesmo que o
    # painel consulta. Só se recorre à recolha em curso quando não há arquivo.
    do_arquivo = historico_do_arquivo(caminho, arquivo) if arquivo else None

    if do_arquivo is not None:
        for dia, registo in sorted(do_arquivo.items()):
            serie["dias"] = [d for d in serie["dias"] if d.get("data") != dia]
            serie["dias"].append({"data": dia, "periodo": "dia completo",
                                  "distintas": registo["distintas"],
                                  "areas": registo["areas"]})
        serie["dias"].sort(key=lambda d: d.get("data", ""))
        with open(caminho, "w", encoding="utf-8") as destino:
            json.dump(serie, destino, ensure_ascii=False, indent=1)
        total = sum(v["noticias"] for r in do_arquivo.values()
                    for v in r["areas"].values())
        print(f"série diária: {len(do_arquivo)} dias reconstruídos do arquivo "
              f"({total} notícias) · {len(serie['dias'])} dias em {caminho}")
        return

    do_dia = [l for l in linhas if l[2] and l[2].strftime("%Y-%m-%d") == hoje]
    print(f"série diária: {len(do_dia)} de {len(linhas)} notícias foram publicadas hoje")

    por_area = {}
    for l in do_dia:
        registo = por_area.setdefault(l[0], {
            "noticias": 0, "novas": 0, "fontes": set(),
            "origens": {"nacionais": 0, "lusofonas": 0, "internacionais": 0},
            "palavras": {},
        })
        registo["noticias"] += 1
        registo["novas"] += 1          # mantido por compatibilidade: é o mesmo
        if l[3]:
            registo["fontes"].add(l[3])
        registo["origens"][origem_da_fonte(l[4])] += 1
        for pal in (l[8] if len(l) > 8 and l[8] else []):
            registo["palavras"][pal] = registo["palavras"].get(pal, 0) + 1

    contagem = {"total": 0, "nacionais": 0, "lusofonas": 0, "internacionais": 0}
    for lig, dom in {l[7]: l[4] for l in do_dia if len(l) > 7 and l[7]}.items():
        contagem["total"] += 1
        contagem[origem_da_fonte(dom)] += 1

    serie["dias"].append({
        "data": hoje,
        "periodo": periodo or "sem limite",
        "distintas": contagem,
        "areas": {nome: {
            "noticias": v["noticias"],
            "novas": v["novas"],
            "fontes": len(v["fontes"]),
            "origens": v["origens"],
            "palavras": dict(sorted(v["palavras"].items(), key=lambda x: -x[1])),
        } for nome, v in sorted(por_area.items())},
    })
    serie["dias"].sort(key=lambda d: d["data"])

    with open(caminho, "w", encoding="utf-8") as destino:
        json.dump(serie, destino, ensure_ascii=False, indent=1)
    print(f"série diária atualizada: {len(serie['dias'])} dias em {caminho}")


if __name__ == "__main__":
    principal()
